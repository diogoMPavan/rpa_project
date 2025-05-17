import requests
from bs4 import BeautifulSoup as bsoup
import json
from pathlib import Path
from openpyxl import Workbook

title = ""
resume = ""
time = ""

def get_html():
    global title, resume, time

    url = 'https://g1.globo.com/'
    headers = {'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/42.0.2311.135 Safari/537.36 Edge/12.246"}

    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        print('Error:', r.status_code)
        return None
    
    soup = bsoup(r.text, 'html5lib')
    latest_news = soup.find('div', class_='feed-post bstn-item-shape type-materia')
    #latest_news = soup.find('div', id='c19736c9-ffb5-4e84-9530-ce59a76e3aa7')

    if latest_news != None:
        #Encontra os itens caso tenha
        items = latest_news.find('div', class_='feed-post-body')
        if items != None:

            ul = items.find('ul', class_='bstn-relateditems')
            if ul:
                for items in ul.find_all('div', class_='bstn-fd-relatedtext'):
                    a = items.find('a', class_='gui-color-primary gui-color-hover feed-post-body-title bstn-relatedtext')
                    if a:
                        resume += a.get_text() + "; "
        body = latest_news.find('div', class_='feed-post-body-title gui-color-primary gui-color-hover')
        title = body.find('a').text
        metadata = latest_news.find('div', class_='feed-post-metadata')
        time = metadata.find('span', class_='feed-post-datetime').text

def save_to_json():
    global title, resume, time
    data = {
        "titulo": title,
        "data_publicacao": time,
        "resumo": resume
    }

    current_dir = Path(__file__).resolve().parent
    main_dir =current_dir.parent
    json_dir = main_dir / "noticia.json"

    with open(json_dir, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def save_to_xlsx():
    global title, resume, time
    current_dir = Path(__file__).resolve().parent
    main_dir = current_dir.parent
    xlsx_dir = main_dir / "noticia.xlsx"

    wb = Workbook()
    ws = wb.active

    ws['A1'] = "titulo"
    ws['B1'] = "data_publicacao"
    ws['C1'] = "resumo"

    ws.append([title, time, resume])
    wb.save(xlsx_dir)

get_html()
save_to_json()
save_to_xlsx()
